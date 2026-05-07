"""
Export analytics service — generates XLSX reports for tracking data.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from sqlalchemy.ext.asyncio import AsyncSession

from services.tracking_service import (
    get_funnel_stats, get_question_dropoff,
    get_top_sources, get_bad_sources,
    get_referral_conversion_stats,
)
from services.referral_service import get_refs_by_form, get_ref_by_id
from services.client_service import get_clients_paginated
from services.offer_service import get_offers_paginated
from services.form_service import get_forms_paginated


def _new_wb_ws(title: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws, Font, PatternFill, Alignment


def _bold_header(ws, headers: list, Font) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Funnel by source export
# ─────────────────────────────────────────────────────────────────────────────

async def export_funnel_by_source(session: AsyncSession) -> bytes:
    """All referral sources with full funnel metrics in one XLSX."""
    top = await get_top_sources(session, limit=1000, order_by="leads_created")

    wb, ws, Font, _, _ = _new_wb_ws("Воронка по источникам")
    headers = [
        "Источник", "Стартов бота", "Завершений формы", "Лидов",
        "Дублей", "Qualified", "Approved",
        "Start→Complete %", "Lead→Qualified %", "Lead→Approved %",
    ]
    _bold_header(ws, headers, Font)

    for r in top:
        ws.append([
            r["ref_name"],
            r["bot_starts"],
            r["form_completions"],
            r["leads_created"],
            r["duplicates"],
            r["qualified"],
            r["approved"],
            r["completion_rate"],
            round(r["qualified"] / r["leads_created"] * 100, 1) if r["leads_created"] else 0,
            r["approve_rate"],
        ])

    _autowidth(ws)
    return _to_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# Drop-off report export
# ─────────────────────────────────────────────────────────────────────────────

async def export_dropoff_report(session: AsyncSession, form_id: int) -> bytes:
    """Per-question drop-off for a single form."""
    dropoff = await get_question_dropoff(session, form_id)

    wb, ws, Font, _, _ = _new_wb_ws("Drop-off по вопросам")
    headers = [
        "Шаг", "Вопрос", "Просмотрели", "Ответили", "Пропустили",
        "Ответили %", "Дропнули %",
    ]
    _bold_header(ws, headers, Font)

    for row in dropoff:
        ws.append([
            row["step"],
            row["question_text"],
            row["viewed"],
            row["answered"],
            row["skipped"],
            row["answer_rate"],
            row["dropoff_pct"],
        ])

    _autowidth(ws)
    return _to_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# Top / bad sources export
# ─────────────────────────────────────────────────────────────────────────────

async def export_top_sources(session: AsyncSession) -> bytes:
    top = await get_top_sources(session, limit=50)
    return _build_source_ranking_xlsx(top, "Топ источники")


async def export_bad_sources(session: AsyncSession) -> bytes:
    bad = await get_bad_sources(session, limit=50)
    return _build_source_ranking_xlsx(bad, "Плохие источники")


def _build_source_ranking_xlsx(data: List[Dict[str, Any]], title: str) -> bytes:
    wb, ws, Font, _, _ = _new_wb_ws(title)
    headers = [
        "Источник", "Стартов бота", "Завершений", "Лидов",
        "Qualified", "Approved", "Completion %", "Approve %",
    ]
    _bold_header(ws, headers, Font)
    for r in data:
        ws.append([
            r["ref_name"],
            r["bot_starts"],
            r["form_completions"],
            r["leads_created"],
            r["qualified"],
            r["approved"],
            r["completion_rate"],
            r["approve_rate"],
        ])
    _autowidth(ws)
    return _to_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# Full analytics bundle
# ─────────────────────────────────────────────────────────────────────────────

async def export_full_analytics(session: AsyncSession) -> bytes:
    """Multi-sheet XLSX: funnel by source, top, bad sources."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    top = await get_top_sources(session, limit=500, order_by="leads_created")
    bad = await get_bad_sources(session, limit=100)

    wb = Workbook()

    # Sheet 1: Funnel by source
    ws1 = wb.active
    ws1.title = "Воронка"
    headers1 = [
        "Источник", "Стартов", "Завершений", "Лидов", "Дублей",
        "Qualified", "Approved", "Completion %", "Approve %",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for r in top:
        ws1.append([
            r["ref_name"], r["bot_starts"], r["form_completions"],
            r["leads_created"], r["duplicates"], r["qualified"], r["approved"],
            r["completion_rate"], r["approve_rate"],
        ])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 50
        )

    # Sheet 2: Bad sources
    ws2 = wb.create_sheet("Плохие источники")
    ws2.append(headers1)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in bad:
        ws2.append([
            r["ref_name"], r["bot_starts"], r["form_completions"],
            r["leads_created"], r["duplicates"], r["qualified"], r["approved"],
            r["completion_rate"], r["approve_rate"],
        ])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 50
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
